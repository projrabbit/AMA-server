from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta, timezone

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.repositories import report_repository
from app.schemas.report import (
    ActiveLocationItem,
    AttendanceReportData,
    DashboardSummaryData,
    RealtimeLocationItem,
    ReportDayDetail,
    ReportEmployeeSummary,
    ReportSummary,
)

_TZ = timezone(timedelta(hours=7))


# ── Dashboard ──────────────────────────────────────────────────────────────────

def get_dashboard_summary(db: Session, target_date: date) -> DashboardSummaryData:
    total_employees = report_repository.get_active_employee_count(db)
    shift_holder_count = report_repository.get_shift_holder_count(db)
    stats = report_repository.get_dashboard_checkin_stats(db, target_date)
    fraud_alerts = report_repository.get_fraud_alert_count(db, target_date)
    checked_in_shift_holders = report_repository.get_checked_in_shift_holders_count(db, target_date)
    active_records = report_repository.get_active_locations(db, target_date)

    checked_in_today = stats["checked_in_today"]
    on_time_count = stats["on_time_count"]
    absent_count = max(0, shift_holder_count - checked_in_shift_holders)
    on_time_rate = round(on_time_count / total_employees * 100, 2) if total_employees > 0 else 0.0

    return DashboardSummaryData(
        date=target_date,
        total_employees=total_employees,
        checked_in_today=checked_in_today,
        on_time_count=on_time_count,
        late_count=stats["late_count"],
        early_leave_count=stats["early_leave_count"],
        absent_count=absent_count,
        fraud_alerts_today=fraud_alerts,
        on_time_rate=on_time_rate,
        active_locations=[_build_active_location(r) for r in active_records],
    )


def _build_active_location(record) -> ActiveLocationItem:
    geo = record.geofence_rule
    cell = geo.cell_space if geo else None
    floor_ = cell.floor if cell else None
    building = cell.building if cell else None

    return ActiveLocationItem(
        employee_id=record.employee_id,
        full_name=record.employee.full_name,
        department_name=record.employee.department.name,
        latitude=float(record.latitude),
        longitude=float(record.longitude),
        altitude=float(record.altitude) if record.altitude is not None else None,
        building_id=building.building_id if building else None,
        building_name=building.name if building else None,
        floor_id=floor_.floor_id if floor_ else None,
        floor_name=floor_.floor_name if floor_ else None,
        last_checkin_at=record.timestamp.astimezone(_TZ),
    )


# ── Realtime ───────────────────────────────────────────────────────────────────

def get_realtime_locations(
    db: Session,
    *,
    building_id: int | None = None,
    floor_id: int | None = None,
    department_id: int | None = None,
) -> list[RealtimeLocationItem]:
    today = datetime.now(_TZ).date()
    records = report_repository.get_realtime_locations(
        db, today, building_id=building_id, floor_id=floor_id, department_id=department_id
    )
    return [_build_realtime_location(r) for r in records]


def _build_realtime_location(record) -> RealtimeLocationItem:
    geo = record.geofence_rule
    cell = geo.cell_space if geo else None
    floor_ = cell.floor if cell else None
    building = cell.building if cell else None

    return RealtimeLocationItem(
        employee_id=record.employee_id,
        full_name=record.employee.full_name,
        department_id=record.employee.department_id,
        department_name=record.employee.department.name,
        record_id=record.record_id,
        latitude=float(record.latitude),
        longitude=float(record.longitude),
        altitude=float(record.altitude) if record.altitude is not None else None,
        gps_accuracy=float(record.gps_accuracy) if record.gps_accuracy is not None else None,
        building_id=building.building_id if building else None,
        building_name=building.name if building else None,
        floor_id=floor_.floor_id if floor_ else None,
        floor_name=floor_.floor_name if floor_ else None,
        arcgis_layer_id=cell.arcgis_layer_id if cell else None,
        checked_in_at=record.timestamp.astimezone(_TZ),
    )


# ── Attendance report ──────────────────────────────────────────────────────────

def get_attendance_report(
    db: Session,
    from_date: date,
    to_date: date,
    *,
    department_id: int | None = None,
    employee_id: int | None = None,
) -> AttendanceReportData:
    records = report_repository.get_report_records(
        db, from_date, to_date, department_id=department_id, employee_id=employee_id
    )
    range_days = (to_date - from_date).days + 1

    emp_checkins: dict[int, list] = defaultdict(list)
    emp_checkouts: dict[int, list] = defaultdict(list)
    emp_info: dict[int, object] = {}

    for r in records:
        emp_info[r.employee_id] = r.employee
        if r.type.value == "checkin":
            emp_checkins[r.employee_id].append(r)
        else:
            emp_checkouts[r.employee_id].append(r)

    all_employee_ids = set(emp_checkins) | set(emp_checkouts)

    employee_summaries: list[ReportEmployeeSummary] = []
    all_details: list[ReportDayDetail] = []

    summary_work_days = summary_work_minutes = 0
    summary_late = summary_early_leave = summary_absent = summary_rejected = 0

    _ok = {"approved", "flagged"}

    for eid in sorted(all_employee_ids):
        emp = emp_info[eid]
        checkins = emp_checkins[eid]
        checkouts = emp_checkouts[eid]

        approved_checkins = [r for r in checkins if r.status.value in _ok]
        checkin_days = {r.timestamp.astimezone(_TZ).date() for r in approved_checkins}
        work_days = len(checkin_days)

        total_mins = sum(r.worked_minutes or 0 for r in checkouts if r.status.value in _ok)
        late_cnt = sum(1 for r in checkins if r.is_late)
        early_cnt = sum(1 for r in checkouts if r.is_early_leave)
        rejected_cnt = sum(1 for r in (checkins + checkouts) if r.status.value == "rejected")
        absent_cnt = max(0, range_days - work_days)

        employee_summaries.append(ReportEmployeeSummary(
            employee_id=eid,
            full_name=emp.full_name,
            department_name=emp.department.name,
            work_days=work_days,
            total_work_minutes=total_mins,
            late_count=late_cnt,
            early_leave_count=early_cnt,
            absent_count=absent_cnt,
            rejected_count=rejected_cnt,
        ))

        summary_work_days += work_days
        summary_work_minutes += total_mins
        summary_late += late_cnt
        summary_early_leave += early_cnt
        summary_absent += absent_cnt
        summary_rejected += rejected_cnt

        checkout_by_checkin: dict[int, object] = {
            co.matched_checkin_record_id: co
            for co in checkouts
            if co.matched_checkin_record_id
        }

        for ci in sorted(checkins, key=lambda r: r.timestamp):
            co = checkout_by_checkin.get(ci.record_id)
            ci_local = ci.timestamp.astimezone(_TZ)
            co_local = co.timestamp.astimezone(_TZ) if co else None

            if co:
                day_status = "completed"
            elif ci.status.value == "rejected":
                day_status = "rejected"
            else:
                day_status = ci.status.value

            all_details.append(ReportDayDetail(
                date=ci_local.date(),
                employee_id=eid,
                full_name=emp.full_name,
                department_name=emp.department.name,
                checkin_at=ci_local,
                checkout_at=co_local,
                worked_minutes=co.worked_minutes if co else None,
                is_late=ci.is_late,
                is_early_leave=co.is_early_leave if co else False,
                status=day_status,
            ))

    all_details.sort(key=lambda d: (d.date, d.employee_id))

    return AttendanceReportData(
        range={"from": str(from_date), "to": str(to_date)},
        summary=ReportSummary(
            employee_count=len(all_employee_ids),
            total_work_days=summary_work_days,
            total_work_minutes=summary_work_minutes,
            late_count=summary_late,
            early_leave_count=summary_early_leave,
            absent_count=summary_absent,
            rejected_count=summary_rejected,
        ),
        employees=employee_summaries,
        details=all_details,
    )


# ── Export ─────────────────────────────────────────────────────────────────────

def export_attendance_report(
    db: Session,
    format_: str,
    from_date: date,
    to_date: date,
    *,
    department_id: int | None = None,
    employee_id: int | None = None,
) -> tuple[bytes, str, str]:
    """Returns (file_bytes, filename, content_type)."""
    if format_ not in ("excel", "pdf"):
        raise HTTPException(
            status_code=400,
            detail={"code": "INVALID_EXPORT_FORMAT", "message": "Format must be 'excel' or 'pdf'", "details": {}},
        )

    data = get_attendance_report(db, from_date, to_date, department_id=department_id, employee_id=employee_id)

    if not data.details and not data.employees:
        raise HTTPException(
            status_code=404,
            detail={"code": "NO_REPORT_DATA", "message": "No attendance data found for the given filters", "details": {}},
        )

    date_suffix = f"{from_date}_{to_date}"

    if format_ == "excel":
        content = _generate_excel(data)
        filename = f"attendance_report_{date_suffix}.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _generate_pdf(data, from_date, to_date)
        filename = f"attendance_report_{date_suffix}.pdf"
        content_type = "application/pdf"

    return content, filename, content_type


def _generate_excel(data: AttendanceReportData) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    s = data.summary
    for label, val in [
        ("Employee Count", s.employee_count),
        ("Total Work Days", s.total_work_days),
        ("Total Work Minutes", s.total_work_minutes),
        ("Late Count", s.late_count),
        ("Early Leave Count", s.early_leave_count),
        ("Absent Count", s.absent_count),
        ("Rejected Count", s.rejected_count),
    ]:
        ws.append([label, val])

    ws2 = wb.create_sheet("By Employee")
    ws2.append(["Employee ID", "Full Name", "Department", "Work Days", "Work Minutes",
                "Late", "Early Leave", "Absent", "Rejected"])
    for e in data.employees:
        ws2.append([e.employee_id, e.full_name, e.department_name, e.work_days,
                    e.total_work_minutes, e.late_count, e.early_leave_count,
                    e.absent_count, e.rejected_count])

    ws3 = wb.create_sheet("Details")
    ws3.append(["Date", "Employee ID", "Full Name", "Department", "Check-in",
                "Check-out", "Worked Min", "Late", "Early Leave", "Status"])
    for row in data.details:
        ws3.append([
            str(row.date), row.employee_id, row.full_name, row.department_name,
            str(row.checkin_at) if row.checkin_at else "",
            str(row.checkout_at) if row.checkout_at else "",
            row.worked_minutes if row.worked_minutes is not None else "",
            row.is_late, row.is_early_leave, row.status,
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_pdf(data: AttendanceReportData, from_date: date, to_date: date) -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Attendance Report: {from_date} to {to_date}", styles["Title"]))
    elements.append(Spacer(1, 12))

    s = data.summary
    elements.append(Paragraph("Summary", styles["Heading2"]))
    t = Table([
        ["Metric", "Value"],
        ["Employee Count", str(s.employee_count)],
        ["Total Work Days", str(s.total_work_days)],
        ["Total Work Minutes", str(s.total_work_minutes)],
        ["Late Count", str(s.late_count)],
        ["Early Leave Count", str(s.early_leave_count)],
        ["Absent Count", str(s.absent_count)],
        ["Rejected Count", str(s.rejected_count)],
    ])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    if data.employees:
        elements.append(Paragraph("By Employee", styles["Heading2"]))
        emp_rows = [["ID", "Name", "Department", "Days", "Minutes", "Late", "Absent"]]
        for e in data.employees:
            emp_rows.append([str(e.employee_id), e.full_name, e.department_name,
                             str(e.work_days), str(e.total_work_minutes),
                             str(e.late_count), str(e.absent_count)])
        t2 = Table(emp_rows)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(t2)

    doc.build(elements)
    return buf.getvalue()
